#!/usr/bin/env python3
"""
Builds the sales dashboard as a static HTML file, ready to publish.

Pulls fresh numbers from Bitrix24 (Disk file, via webhook) and the
"Производство" Google Sheet (public CSV export), then renders them into
dashboard_template.html (replacing the __DASHBOARD_DATA__ marker with a JSON
blob the page's own JS reads on load).

Run hourly by the "Update dashboard" GitHub Actions workflow
(.github/workflows/update.yml). Reads the Bitrix24 webhook base URL from the
BITRIX_WEBHOOK environment variable (set as a repo secret in Actions — never
hardcode it here, this file is public). Output: index.html at the repo root,
which GitHub Pages serves directly.

Usage: BITRIX_WEBHOOK=https://.../rest/N/token python3 scripts/build_dashboard.py
Requires: openpyxl  (pip install openpyxl)
"""

import csv
import io
import json
import os
import re
import sys
import urllib.request
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
REPO_ROOT = HERE.parent

# ---- Bitrix24 -----------------------------------------------------------

BITRIX_WEBHOOK = os.environ["BITRIX_WEBHOOK"]
BITRIX_FILE_ID = "453924"  # "1. План-факт (2026).xlsx" on Disk
BITRIX_SHEET_NAME = "План-факт"

MONTHS_RU = [
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]

DEPARTMENTS = [
    {"name": "ГОСТ", "managers": ["Кочураев", "Угольников", "Филаретова"]},
    {"name": "Проектный", "managers": ["Якунин", "Роганин", "Маркова", "Малашкин", "Извекова", "Ерицян"]},
]

# ---- Google Sheets (Производство) ---------------------------------------

PROD_SHEET_ID = "1AlcRl0byrg1HSVDhboZ150FvOi-hVd3wEpIwFqWgRZQ"
PROD_GID = "614419876"
PROD_DATA_START_ROW = 19  # 1-indexed row where the per-date shipment table begins


def http_get_json(url):
    with urllib.request.urlopen(url, timeout=30) as r:
        return json.loads(r.read().decode("utf-8"))


def http_get_bytes(url):
    with urllib.request.urlopen(url, timeout=60) as r:
        return r.read()


def num(v):
    if v is None or v == "" or v == "-" or v == "?":
        return 0.0
    if isinstance(v, str):
        v = v.replace("\xa0", "").replace(" ", "").replace(",", ".")
        try:
            return float(v)
        except ValueError:
            return 0.0
    return float(v)


def fetch_bitrix_plan_fact(current_month_index):
    meta = http_get_json(f"{BITRIX_WEBHOOK}/disk.file.get.json?id={BITRIX_FILE_ID}")
    if "result" not in meta:
        raise RuntimeError(f"disk.file.get failed: {meta}")
    download_url = meta["result"]["DOWNLOAD_URL"]
    updated_time = meta["result"].get("UPDATE_TIME")

    xlsx_bytes = http_get_bytes(download_url)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb[BITRIX_SHEET_NAME]

    month_row = {}
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        if isinstance(v, str) and v.strip() in MONTHS_RU:
            month_row[v.strip()] = row

    months_to_sum = MONTHS_RU[: current_month_index + 1]
    missing = [m for m in months_to_sum if m not in month_row]
    if missing:
        raise RuntimeError(f"Could not find rows for months: {missing}")

    manager_col = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=col).value
        if isinstance(v, str) and v.strip():
            name = v.strip()
            bare = re.sub(r"\s*\(.*?\)\s*", "", name).strip()
            manager_col[name] = col
            manager_col.setdefault(bare, col)

    result = {}
    for dept in DEPARTMENTS:
        dept_out = []
        for mgr in dept["managers"]:
            if mgr not in manager_col:
                raise RuntimeError(f"Manager column not found for '{mgr}'")
            col = manager_col[mgr]
            plan_sum = fact_sum = 0.0
            for m in months_to_sum:
                r = month_row[m]
                plan_sum += num(ws.cell(row=r, column=col).value)
                fact_sum += num(ws.cell(row=r, column=col + 1).value)
            pct = (fact_sum / plan_sum * 100) if plan_sum else 0.0
            dept_out.append({
                "name": mgr,
                "plan": round(plan_sum),
                "fact": round(fact_sum),
                "pct": round(pct, 1),
            })
        result[dept["name"]] = dept_out

    return result, updated_time


def fetch_production():
    url = f"https://docs.google.com/spreadsheets/d/{PROD_SHEET_ID}/export?format=csv&gid={PROD_GID}"
    raw = http_get_bytes(url).decode("utf-8")
    rows = list(csv.reader(io.StringIO(raw)))

    shipments_count = 0
    late_shipments_count = 0
    last_in_work_count = None
    last_in_work_sum = None

    for r in rows[PROD_DATA_START_ROW - 1:]:
        if len(r) > 1:
            b = r[1].strip()
            if b.replace(" ", "").isdigit():
                shipments_count += int(b)
        if len(r) > 6:
            g = r[6].strip()
            if g.replace(" ", "").lstrip("-").isdigit():
                late_shipments_count += int(g)
        if len(r) > 4:
            d = r[3].strip()
            e = r[4].strip()
            if d not in ("", "-") and e not in ("", "-"):
                last_in_work_count = int(num(d))
                last_in_work_sum = num(e)

    return {
        "shipments": {"count": shipments_count},
        "lateShipments": {"count": late_shipments_count},
        "production": {
            "count": last_in_work_count or 0,
            "sum": round(last_in_work_sum or 0),
        },
    }


def build_data():
    moscow_now = datetime.now(timezone(timedelta(hours=3)))
    current_month_index = moscow_now.month - 1

    plan_fact, bitrix_updated = fetch_bitrix_plan_fact(current_month_index)
    prod = fetch_production()

    return {
        "updatedAt": moscow_now.isoformat(),
        "currentMonth": MONTHS_RU[current_month_index],
        "departments": [
            {"name": dept["name"], "managers": plan_fact[dept["name"]]}
            for dept in DEPARTMENTS
        ],
        "shipments": prod["shipments"],
        "lateShipments": prod["lateShipments"],
        "production": prod["production"],
        "sourceUpdatedAt": bitrix_updated,
    }


def main():
    data = build_data()
    template = (HERE / "dashboard_template.html").read_text(encoding="utf-8")
    rendered = template.replace(
        "/*__DASHBOARD_DATA__*/null",
        json.dumps(data, ensure_ascii=False),
    )
    (REPO_ROOT / "index.html").write_text(rendered, encoding="utf-8")
    print("Wrote index.html")
    print(json.dumps(data, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)
